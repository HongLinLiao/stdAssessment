from os import walk
from os.path import join
from openpyxl import load_workbook

from assessment import parse

def root(path):
    # result excel
    excel = load_workbook('./result.xlsx')
    # first sheet
    worksheet = excel.get_sheet_by_name(excel.sheetnames[0])

    max_row = worksheet.max_row + 1

    # 標頭
    worksheet['A1'] = 'Id'
    worksheet['B1'] = 'School'
    worksheet['C1'] = 'Name'
    worksheet['D1'] = 'IsOk'
    worksheet['E1'] = 'Message'
    worksheet['F1'] = '高一上學期班級排名百分比'
    worksheet['G1'] = '高一上學期類組排名百分比'
    worksheet['H1'] = '高一上學期年級排名百分比'
    worksheet['I1'] = '高一下學期班級排名百分比'
    worksheet['J1'] = '高一下學期類組排名百分比'
    worksheet['K1'] = '高一下學期年級排名百分比'
    worksheet['L1'] = '高二上學期班級排名百分比'
    worksheet['M1'] = '高二上學期類組排名百分比'
    worksheet['N1'] = '高二上學期年級排名百分比'
    worksheet['O1'] = '高二下學期班級排名百分比'
    worksheet['P1'] = '高二下學期類組排名百分比'
    worksheet['Q1'] = '高二下學期年級排名百分比'
    worksheet['R1'] = '高三上學期班級排名百分比'
    worksheet['S1'] = '高三上學期類組排名百分比'
    worksheet['T1'] = '高三上學期年級排名百分比'

    # pdf 數量
    count = 0;

    # 所有檔案的絕對路徑
    for root, dirs, files in walk(path):
        for f in files:
            fullpath = join(root, f)
            # 篩掉不是pdf的file
            if(fullpath.find('.pdf') != -1):
                # call parse
                result = parse(fullpath)
                worksheet['A'+str(max_row+count)] = result[0].id
                worksheet['B'+str(max_row+count)] = result[0].school
                worksheet['C'+str(max_row+count)] = result[0].name
                if(result[0].isOk == False):
                    worksheet['D'+str(max_row+count)] = 'x'
                    worksheet['E'+str(max_row+count)] = result[0].message
                else:
                    worksheet['F'+str(max_row+count)] = result[1].firstＵpClassPercentage
                    worksheet['G'+str(max_row+count)] = result[1].firstUpCategoryPercentage
                    worksheet['H'+str(max_row+count)] = result[1].firstUpAllPercentage
                    worksheet['I'+str(max_row+count)] = result[1].firstDownClassPercentage
                    worksheet['J'+str(max_row+count)] = result[1].firstDownCategoryPercentage
                    worksheet['K'+str(max_row+count)] = result[1].firstDownAllPercentage
                    worksheet['L'+str(max_row+count)] = result[1].secondＵpClassPercentage
                    worksheet['M'+str(max_row+count)] = result[1].secondUpCategoryPercentage
                    worksheet['N'+str(max_row+count)] = result[1].secondUpAllPercentage
                    worksheet['O'+str(max_row+count)] = result[1].secondDownClassPercentage
                    worksheet['P'+str(max_row+count)] = result[1].secondDownCategoryPercentage
                    worksheet['Q'+str(max_row+count)] = result[1].secondDownAllPercentage
                    worksheet['R'+str(max_row+count)] = result[1].thirdＵpClassPercentage
                    worksheet['S'+str(max_row+count)] = result[1].thirdUpCategoryPercentage
                    worksheet['T'+str(max_row+count)] = result[1].thirdUpAllPercentage

                print("%s,%s,%s,%s,%s" % (result[0].id,result[0].school,result[0].name,result[0].isOk,result[0].message))

                count = count + 1
    excel.save('result.xlsx')
    print('End!')