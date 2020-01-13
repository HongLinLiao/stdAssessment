from openpyxl import load_workbook, Workbook

class XLSX():

    # init workbook : private
    def __init__(self, path):

        try:       
            self.workbook = load_workbook(path)
        except Exception as e:
            try:
                wb = Workbook()
                wb.save(path)
                self.workbook = load_workbook(path)
            except Exception as e:
                print('創建輸出檔失敗！')

        self.filename = path.split('/')[len(path.split('/'))-1]

    # return sheet by name
    def getSheetByName(self, sheetName):
        return self.workbook.get_sheet_by_name(sheetName)

    # return sheet max row index
    def getSheetMaxRow(self, sheet):
        return sheet.max_row

    # write data on sheet location
    def writeDataByLocation(self, sheet, location, data):
        sheet[location] = data
    
    def saveWorkbook(self):
        self.workbook.save(self.filename)
